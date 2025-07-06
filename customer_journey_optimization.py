import pandas as pd
import matplotlib.pyplot as plt
import mysql.connector
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timedelta

# Example MySQL connection configuration
DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': 'Aryan14mysql',
    'database': 'customer_data',
    'port': 3306
}

# 1. Connect to MySQL database
conn = mysql.connector.connect(**DB_CONFIG)
cursor = conn.cursor(dictionary=True)

# 2. Extract 12 months of data from each table
end_date = datetime.today()
start_date = end_date - timedelta(days=365)

# Query for deliveries
deliveries_query = f'''
    SELECT order_id, customer_id, delivery_date, promised_date
    FROM deliveries
    WHERE delivery_date BETWEEN '{start_date.date()}' AND '{end_date.date()}'
'''
deliveries = pd.read_sql(deliveries_query, conn)

# Query for complaints
complaints_query = f'''
    SELECT complaint_id, customer_id, complaint_date, category
    FROM complaints
    WHERE complaint_date BETWEEN '{start_date.date()}' AND '{end_date.date()}'
'''
complaints = pd.read_sql(complaints_query, conn)

# Query for escalations
escalations_query = f'''
    SELECT escalation_id, customer_id, escalation_date, resolution_time_hours
    FROM escalations
    WHERE escalation_date BETWEEN '{start_date.date()}' AND '{end_date.date()}'
'''
escalations = pd.read_sql(escalations_query, conn)

# Close the cursor and connection
cursor.close()
conn.close()

# 3. Clean and transform data
# Convert date fields to datetime
for df, date_cols in [
    (deliveries, ['delivery_date', 'promised_date']),
    (complaints, ['complaint_date']),
    (escalations, ['escalation_date'])
]:
    for col in date_cols:
        df[col] = pd.to_datetime(df[col])

# Calculate delivery delays in days
deliveries['delivery_delay_days'] = (deliveries['delivery_date'] - deliveries['promised_date']).dt.days

# 4. Merge the tables on customer_id for combined analysis
# Merge deliveries and complaints
merged = pd.merge(deliveries, complaints, on='customer_id', how='outer', suffixes=('_delivery', '_complaint'))
# Merge with escalations
merged = pd.merge(merged, escalations, on='customer_id', how='outer', suffixes=('', '_escalation'))

# 5. Perform analysis
# Pareto chart of complaints by category
complaint_counts = complaints['category'].value_counts().sort_values(ascending=False)
cum_percentage = complaint_counts.cumsum() / complaint_counts.sum() * 100

plt.figure(figsize=(10,6))
complaint_counts.plot(kind='bar', color='skyblue')
plt.ylabel('Number of Complaints')
plt.title('Pareto Chart of Complaints by Category')
plt.twinx()
cum_percentage.plot(color='red', marker='o', linewidth=2)
plt.ylabel('Cumulative Percentage (%)')
plt.grid(axis='y')
plt.tight_layout()
plt.savefig('pareto_complaints.png')
plt.close()

# Calculate KPIs
kpi_summary = {
    'Total Deliveries': len(deliveries),
    'Total Complaints': len(complaints),
    'Total Escalations': len(escalations),
    'Average Delivery Delay (days)': deliveries['delivery_delay_days'].mean(),
    'Min Delivery Delay (days)': deliveries['delivery_delay_days'].min(),
    'Max Delivery Delay (days)': deliveries['delivery_delay_days'].max(),
}
kpi_df = pd.DataFrame(list(kpi_summary.items()), columns=['KPI', 'Value'])

# 6. Generate an Excel dashboard
wb = Workbook()

# Sheet 1: Raw Deliveries data
ws1 = wb.active
ws1.title = 'Deliveries'
for r in dataframe_to_rows(deliveries, index=False, header=True):
    ws1.append(r)

# Sheet 2: Raw Complaints data
ws2 = wb.create_sheet('Complaints')
for r in dataframe_to_rows(complaints, index=False, header=True):
    ws2.append(r)

# Sheet 3: Raw Escalations data
ws3 = wb.create_sheet('Escalations')
for r in dataframe_to_rows(escalations, index=False, header=True):
    ws3.append(r)

# Sheet 4: Merged Data
ws4 = wb.create_sheet('Merged Data')
for r in dataframe_to_rows(merged, index=False, header=True):
    ws4.append(r)

# Sheet 5: KPI Summary
ws5 = wb.create_sheet('KPI Summary')
for r in dataframe_to_rows(kpi_df, index=False, header=True):
    ws5.append(r)

# Save the Excel file
wb.save('customer_journey_dashboard.xlsx')

# 7. Pareto chart already saved as 'pareto_complaints.png'

# 8. Code is commented for clarity

print('Dashboard and Pareto chart generated successfully.') 